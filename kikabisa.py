import os
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
import PyPDF2
import docx
import openpyxl
import numpy as np
import subprocess
from collections import Counter  # Add this line to import Counter

# Pembuatan stemmer di awal untuk efisiensi
factory = StemmerFactory()
stemmer = factory.create_stemmer()

# Fungsi untuk membaca stopwords dari file
def load_stopwords(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8') as file:
            stopwords = set(file.read().splitlines())  # Membaca dan memisahkan baris
        return stopwords
    except Exception as e:
        print(f"Error loading stopwords: {e}")
        return set()

# Memuat stopwords dari file
stopwords_file = 'stopwordbahasa.csv' # Ganti dengan path yang sesuai
stopwords = load_stopwords(stopwords_file)


def case_folding(teks):
    return teks.lower()


def tokenisasi(teks):
    return teks.split()


def filtering(token):
    return [kata for kata in token if kata not in stopwords]  # Menggunakan stopwords dari file


def stemming(token):
    return [stemmer.stem(kata) for kata in token]


def baca_file(filepath):
    try:
        if filepath.endswith('.txt'):
            with open(filepath, 'r', encoding='utf-8') as file:
                return file.read()
        elif filepath.endswith('.pdf'):
            pdf_reader = PyPDF2.PdfReader(filepath)
            return ''.join([page.extract_text() for page in pdf_reader.pages])
        elif filepath.endswith('.docx'):
            doc = docx.Document(filepath)
            return '\n'.join([paragraf.text for paragraf in doc.paragraphs])
        elif filepath.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(filepath)
            teks = ''
            for sheet in workbook.sheetnames:
                sheet_data = workbook[sheet]
                for row in sheet_data.iter_rows(values_only=True):
                    teks += ' '.join(map(str, row)) + '\n'
            return teks
        else:
            return ''
    except Exception as e:
        return f"Error reading file {filepath}: {e}"


def preprocessing(teks):
    dilipat = case_folding(teks)
    token = tokenisasi(dilipat)
    difilter = filtering(token)
    distem = stemming(difilter)
    return dilipat, token, difilter, distem


def hitung_gvsm(kueri, dokumen):
    vocabulary = list(set(kueri + [kata for doc in dokumen for kata in doc]))
    term_index = {term: i for i, term in enumerate(vocabulary)}

    correlation_matrix = np.identity(len(vocabulary))

    def vectorize(text):
        vec = [0] * len(vocabulary)
        for word in text:
            if word in term_index:
                vec[term_index[word]] += 1
        return vec

    kueri_vec = np.dot(vectorize(kueri), correlation_matrix)
    dokumen_vecs = [np.dot(vectorize(doc), correlation_matrix) for doc in dokumen]

    def cosine_similarity(vec1, vec2):
        dot_product = np.dot(vec1, vec2)
        magnitude1 = np.linalg.norm(vec1)
        magnitude2 = np.linalg.norm(vec2)
        if magnitude1 == 0 or magnitude2 == 0:
            return 0
        return dot_product / (magnitude1 * magnitude2)

    return [cosine_similarity(kueri_vec, doc_vec) for doc_vec in dokumen_vecs]


def proses_kueri():
    folder = folder_var.get()
    kueri = kueri_var.get()
    hasil_tabel.delete(*hasil_tabel.get_children())

    if not folder or not kueri:
        hasil_box.insert(tk.END, 'Silakan pilih folder dan masukkan kueri.\n')
        return

    files = daftar_file.get(0, tk.END)
    dokumen = []
    nama_file = []

    for file in files:
        filepath = os.path.join(folder, file)
        teks = baca_file(filepath)
        _, _, _, teks_distem = preprocessing(teks)

        dokumen.append(teks_distem)
        nama_file.append(file)

    kueri_distem = preprocessing(kueri)[3]
    hasil_box.insert(tk.END, f"Kueri Setelah Preprocessing: {' '.join(kueri_distem)}\n\n")

    kesamaan = hitung_gvsm(kueri_distem, dokumen)

    hasil_box.insert(tk.END, "Hasil Kesamaan:\n")
    for file, similarity in sorted(zip(nama_file, kesamaan), key=lambda x: x[1], reverse=True):
        hasil_tabel.insert('', 'end', values=(file, f"{similarity:.4f}"))

    hasil_box.insert(tk.END, "Pilih dokumen untuk melihat detail.\n")


def pilih_folder():
    folder = filedialog.askdirectory()
    folder_var.set(folder)
    if folder:
        try:
            files = os.listdir(folder)
            daftar_file.delete(0, tk.END)
            for file in files:
                if file.endswith(('.txt', '.pdf', '.docx', '.xlsx')):
                    daftar_file.insert(tk.END, file)
            if not files:
                hasil_box.insert(tk.END, "Tidak ada file yang cocok ditemukan di folder.\n")
        except Exception as e:
            hasil_box.insert(tk.END, f"Error membaca folder: {e}\n")


def tampilkan_detail(event):
    selected_item = hasil_tabel.selection()
    if not selected_item:
        return

    item = hasil_tabel.item(selected_item)
    file_name = item['values'][0]
    file_path = os.path.join(folder_var.get(), file_name)
    teks = baca_file(file_path)
    dilipat, token, difilter, distem = preprocessing(teks)

    detail_window = tk.Toplevel(root)
    detail_window.title(f"Detail Dokumen: {file_name}")

    text_area = scrolledtext.ScrolledText(detail_window, width=80, height=20)
    text_area.pack(padx=10, pady=10)
    text_area.insert(tk.END, teks)

    button_frame = tk.Frame(detail_window)
    button_frame.pack(pady=10)

    # def open_file():
    #     try:
    #         subprocess.run(["open" if os.name == "posix" else "start", file_path], shell=True)
    #     except Exception as e:
    #         text_area.insert(tk.END, f"\nError opening file: {e}")

    def open_file():
        try:
            result = subprocess.run(["open" if os.name == "posix" else "start", file_path], shell=True,
                                    capture_output=True, text=True)
            if result.returncode != 0:
                text_area.insert(tk.END, f"Error opening file: {result.stderr}\n")
            else:
                text_area.insert(tk.END, "File opened successfully.\n")
        except Exception as e:
            text_area.insert(tk.END, f"Error opening file: {e}\n")

    def show_kata_dasar():
        dasar = stemming(token)  # Ambil kata dasar dari hasil stemming
        dasar_counter = Counter(dasar)  # Hitung frekuensi setiap kata dasar
        text_area.insert(tk.END, "\n\nKata Dasar dan Frekuensinya:\n")

        # Menampilkan setiap kata dasar dan jumlah kemunculannya
        for kata, jumlah in dasar_counter.items():
            text_area.insert(tk.END, f"{kata}: {jumlah}\n")

    def show_tokenizing():
        text_area.insert(tk.END, f"\n\nTokenizing:\n{' '.join(token)}")

    def show_filtering():
        text_area.insert(tk.END, f"\n\nFiltering:\n{' '.join(difilter)}")

    def show_stemming():
        text_area.insert(tk.END, f"\n\nStemming:\n{' '.join(distem)}")

    tk.Button(button_frame, text="Open", command=open_file).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Kata Dasar", command=show_kata_dasar).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Read Document",
              command=lambda: text_area.insert(tk.END, f"\n\nIsi Dokumen:\n{teks}")).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Tokenizing", command=show_tokenizing).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Filtration", command=show_filtering).pack(side=tk.LEFT, padx=5)
    tk.Button(button_frame, text="Stemming", command=show_stemming).pack(side=tk.LEFT, padx=5)


root = tk.Tk()
root.title("Sistem Temu Kembali Dokumen dengan GVSM")

folder_var = tk.StringVar()
kueri_var = tk.StringVar()

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

label_folder = tk.Label(frame, text="Folder:")
label_folder.grid(row=0, column=0, sticky="w")

entry_folder = tk.Entry(frame, textvariable=folder_var, width=50)
entry_folder.grid(row=0, column=1, padx=5)

tombol_folder = tk.Button(frame, text="Pilih Folder", command=pilih_folder)
tombol_folder.grid(row=0, column=2, padx=5)

daftar_file = tk.Listbox(frame, width=50, height=10)
daftar_file.grid(row=1, column=0, columnspan=3, pady=5)

label_kueri = tk.Label(frame, text="Kueri:")
label_kueri.grid(row=2, column=0, sticky="w")

entry_kueri = tk.Entry(frame, textvariable=kueri_var, width=50)
entry_kueri.grid(row=2, column=1, padx=5)

tombol_cari = tk.Button(frame, text="Cari", command=proses_kueri)
tombol_cari.grid(row=2, column=2, padx=5)

kolom = ("Nama File", "Kesamaan")
hasil_tabel = ttk.Treeview(frame, columns=kolom, show="headings", height=8)
hasil_tabel.grid(row=3, column=0, columnspan=3, pady=5)
hasil_tabel.heading("Nama File", text="Nama File")
hasil_tabel.heading("Kesamaan", text="Kesamaan")

hasil_tabel.bind("<Double-1>", tampilkan_detail)

hasil_box = scrolledtext.ScrolledText(frame, width=80, height=20)
hasil_box.grid(row=4, column=0, columnspan=3, pady=10)

root.mainloop()
